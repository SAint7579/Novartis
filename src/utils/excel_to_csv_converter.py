"""
Utility script to convert large Excel files to CSV format efficiently.
Handles files that are too large for pandas to load into memory.
"""

import os
import csv
from pathlib import Path
from typing import Optional, List
from openpyxl import load_workbook
from tqdm import tqdm


def convert_excel_to_csv(
    excel_path: str,
    output_dir: Optional[str] = None,
    sheet_name: Optional[str] = None,
    chunk_size: int = 1000,
    verbose: bool = True
) -> List[str]:
    """
    Convert Excel file to CSV format efficiently using streaming.
    
    Parameters:
    -----------
    excel_path : str
        Path to the Excel file (.xlsx)
    output_dir : str, optional
        Directory to save CSV files. If None, saves in same directory as Excel file
    sheet_name : str, optional
        Specific sheet name to convert. If None, converts all sheets
    chunk_size : int
        Number of rows to process at a time (for progress display)
    verbose : bool
        Whether to print progress information
    
    Returns:
    --------
    List[str]
        List of paths to created CSV files
    """
    
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    
    # Set output directory
    if output_dir is None:
        output_dir = excel_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
    
    if verbose:
        print(f"Loading workbook: {excel_path.name}")
        print(f"File size: {excel_path.stat().st_size / (1024**2):.2f} MB")
    
    # Load workbook in read-only mode for memory efficiency
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    
    csv_files = []
    
    # Get sheets to process
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        sheets_to_process = [sheet_name]
    else:
        sheets_to_process = workbook.sheetnames
    
    if verbose:
        print(f"Found {len(workbook.sheetnames)} sheet(s): {workbook.sheetnames}")
        print(f"Converting {len(sheets_to_process)} sheet(s)")
    
    # Process each sheet
    for sheet_name in sheets_to_process:
        sheet = workbook[sheet_name]
        
        # Create output filename
        base_name = excel_path.stem
        if len(sheets_to_process) > 1:
            csv_filename = f"{base_name}_{sheet_name}.csv"
        else:
            csv_filename = f"{base_name}.csv"
        
        csv_path = output_dir / csv_filename
        
        if verbose:
            print(f"\nConverting sheet: '{sheet_name}' -> {csv_filename}")
        
        # Write to CSV
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            csv_writer = csv.writer(csvfile)
            
            row_count = 0
            
            # Use tqdm for progress bar if verbose
            if verbose:
                # Get approximate row count (may not be exact for read-only mode)
                try:
                    max_row = sheet.max_row
                except:
                    max_row = None
                
                pbar = tqdm(total=max_row, desc=f"Processing rows", unit="rows")
            
            # Stream rows from Excel
            for row in sheet.iter_rows(values_only=True):
                # Convert None values to empty strings and handle data types
                cleaned_row = []
                for cell in row:
                    if cell is None:
                        cleaned_row.append('')
                    else:
                        cleaned_row.append(str(cell))
                
                csv_writer.writerow(cleaned_row)
                row_count += 1
                
                if verbose and row_count % chunk_size == 0:
                    pbar.update(chunk_size)
            
            if verbose:
                pbar.update(row_count % chunk_size)  # Update remaining rows
                pbar.close()
        
        csv_files.append(str(csv_path))
        
        if verbose:
            print(f"✓ Saved: {csv_path}")
            print(f"  Rows processed: {row_count:,}")
            print(f"  File size: {csv_path.stat().st_size / (1024**2):.2f} MB")
    
    workbook.close()
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"Conversion complete! Created {len(csv_files)} CSV file(s)")
    
    return csv_files


def convert_multiple_excel_files(
    excel_dir: str,
    output_dir: Optional[str] = None,
    pattern: str = "*.xlsx"
) -> List[str]:
    """
    Convert multiple Excel files in a directory to CSV.
    
    Parameters:
    -----------
    excel_dir : str
        Directory containing Excel files
    output_dir : str, optional
        Directory to save CSV files
    pattern : str
        File pattern to match (default: "*.xlsx")
    
    Returns:
    --------
    List[str]
        List of paths to all created CSV files
    """
    excel_dir = Path(excel_dir)
    excel_files = list(excel_dir.glob(pattern))
    
    if not excel_files:
        print(f"No files matching '{pattern}' found in {excel_dir}")
        return []
    
    print(f"Found {len(excel_files)} Excel file(s) to convert")
    
    all_csv_files = []
    
    for excel_file in excel_files:
        print(f"\n{'='*60}")
        csv_files = convert_excel_to_csv(excel_file, output_dir=output_dir)
        all_csv_files.extend(csv_files)
    
    return all_csv_files


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Convert large Excel files to CSV format efficiently"
    )
    parser.add_argument(
        "excel_path",
        help="Path to Excel file or directory containing Excel files"
    )
    parser.add_argument(
        "-o", "--output",
        help="Output directory for CSV files (default: same as input)",
        default=None
    )
    parser.add_argument(
        "-s", "--sheet",
        help="Specific sheet name to convert (default: all sheets)",
        default=None
    )
    parser.add_argument(
        "-c", "--chunk-size",
        help="Chunk size for processing (default: 1000)",
        type=int,
        default=1000
    )
    parser.add_argument(
        "-q", "--quiet",
        help="Suppress progress output",
        action="store_true"
    )
    
    args = parser.parse_args()
    
    input_path = Path(args.excel_path)
    
    if input_path.is_file():
        # Convert single file
        convert_excel_to_csv(
            excel_path=str(input_path),
            output_dir=args.output,
            sheet_name=args.sheet,
            chunk_size=args.chunk_size,
            verbose=not args.quiet
        )
    elif input_path.is_dir():
        # Convert all Excel files in directory
        convert_multiple_excel_files(
            excel_dir=str(input_path),
            output_dir=args.output
        )
    else:
        print(f"Error: '{input_path}' is not a valid file or directory")

